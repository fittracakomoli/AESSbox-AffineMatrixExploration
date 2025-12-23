import json
import csv
import io
import re
from typing import List, Optional, TypedDict
from fastapi import UploadFile, HTTPException
from openpyxl import load_workbook

def _parse_sbox_token(token) -> int:
    if isinstance(token, int):
        return token
    if isinstance(token, float) and token.is_integer():
        return int(token)
    if not isinstance(token, str):
        raise ValueError("Token bukan string atau angka.")
    text = token.strip()
    if not text:
        raise ValueError("Token kosong.")
    if text.lower().startswith("0x"):
        return int(text, 16)
    if re.search(r"[a-fA-F]", text):
        return int(text, 16)
    return int(text, 10)


class ParsedSBox(TypedDict):
    sbox: List[int]
    affine_matrix: Optional[List[List[int]]]
    affine_vector: Optional[List[int]]

def _parse_affine_matrix(matrix_value) -> Optional[List[List[int]]]:
    if matrix_value is None:
        return None
    if not isinstance(matrix_value, list) or len(matrix_value) != 8:
        raise ValueError("Format affine_matrix harus list 8x8.")
    parsed_rows = []
    for row in matrix_value:
        if not isinstance(row, list) or len(row) != 8:
            raise ValueError("Format affine_matrix harus list 8x8.")
        parsed_row = [_parse_sbox_token(item) for item in row]
        if any(val not in (0, 1) for val in parsed_row):
            raise ValueError("Nilai affine_matrix harus 0 atau 1.")
        parsed_rows.append(parsed_row)
    return parsed_rows

def _parse_affine_vector(vector_value) -> Optional[List[int]]:
    if vector_value is None:
        return None
    if not isinstance(vector_value, list) or len(vector_value) != 8:
        raise ValueError("Format affine_vector harus list dengan 8 elemen.")
    parsed_vector = [_parse_sbox_token(item) for item in vector_value]
    if any(val not in (0, 1) for val in parsed_vector):
        raise ValueError("Nilai affine_vector harus 0 atau 1.")
    return parsed_vector

async def parse_uploaded_sbox(file: UploadFile) -> ParsedSBox:
    """
    Membaca file upload (JSON/CSV/TXT/XLSX) dan mengonversinya menjadi List[int].
    Mendukung format Desimal dan Hex (mis. "5B" atau "0x5B").
    """
    filename = file.filename.lower()
    content = await file.read()

    sbox_data = []
    affine_matrix = None
    affine_vector = None

    try:
        # --- KASUS 1: File Excel (.xlsx) ---
        if filename.endswith(".xlsx"):
            wb = load_workbook(filename=io.BytesIO(content), data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    try:
                        value = _parse_sbox_token(cell)
                    except (TypeError, ValueError):
                        continue
                    sbox_data.append(value)

        # Decode bytes ke string untuk format text
        else:
            try:
                text_content = content.decode("utf-8")
            except UnicodeDecodeError:
                raise HTTPException(400, "File harus berupa text (UTF-8).")

        # --- KASUS 1: File JSON ---
            if filename.endswith(".json"):
                data = json.loads(text_content)
                # Handle format { "sbox": [...] } atau langsung [...]
                if isinstance(data, dict) and "sbox" in data:
                    sbox_data = [_parse_sbox_token(item) for item in data["sbox"]]
                    affine_matrix = _parse_affine_matrix(data.get("affine_matrix", data.get("matrix")))
                    affine_vector = _parse_affine_vector(data.get("affine_vector", data.get("vector")))
                elif isinstance(data, list):
                    sbox_data = [_parse_sbox_token(item) for item in data]
                else:
                    raise ValueError("JSON harus berisi array atau object dengan key 'sbox'.")

            # --- KASUS 2: File CSV (.csv) ---
            elif filename.endswith(".csv"):
                f = io.StringIO(text_content)
                reader = csv.reader(f)
                for row in reader:
                    for item in row:
                        # Bersihkan whitespace dan parse int
                        clean_item = item.strip()
                        if clean_item:
                            try:
                                sbox_data.append(_parse_sbox_token(clean_item))
                            except ValueError:
                                continue # Skip header atau text non-angka

            # --- KASUS 3: File Text (.txt) ---
            elif filename.endswith(".txt"):
                # TXT mungkin dipisah spasi, tab, atau enter.
                # Ambil token sederhana (angka desimal atau hex) dari teks.
                tokens = re.split(r'[\s,;]+', text_content.strip())
                sbox_data = []
                for token in tokens:
                    if not token:
                        continue
                    try:
                        sbox_data.append(_parse_sbox_token(token))
                    except ValueError:
                        continue

            else:
                raise HTTPException(400, "Format file tidak didukung. Gunakan .json, .csv, .txt, atau .xlsx")

    except json.JSONDecodeError:
        raise HTTPException(400, "File JSON rusak/tidak valid.")
    except Exception as e:
        raise HTTPException(400, f"Gagal membaca file: {str(e)}")

    # --- VALIDASI PENTING ---
    # S-box AES harus tepat 256 elemen
    if len(sbox_data) != 256:
        raise HTTPException(
            status_code=400, 
            detail=f"Jumlah data tidak valid. Ditemukan {len(sbox_data)} angka, seharusnya tepat 256."
        )

    # Validasi Range (Harus 0-255)
    if any(x < 0 or x > 255 for x in sbox_data):
        raise HTTPException(400, "Nilai S-box harus berada dalam rentang 0-255.")

    return {
        "sbox": sbox_data,
        "affine_matrix": affine_matrix,
        "affine_vector": affine_vector
    }

def format_sbox_as_csv(sbox: List[int]) -> io.StringIO:
    """
    Mengubah S-box menjadi format CSV Grid 16x16 dalam bentuk HEX (2 digit).
    """
    output = io.StringIO()
    writer = csv.writer(output)
    
    for i in range(0, 256, 16):
        row_ints = sbox[i : i+16]
        row_hex = [f"{val:02X}" for val in row_ints]
        writer.writerow(row_hex)
        
    output.seek(0)
    return output

def format_sbox_as_txt(sbox: List[int]) -> io.StringIO:
    """
    Mengubah S-box menjadi format TXT Grid 16x16 dalam bentuk HEX (2 digit).
    Dipisahkan dengan spasi.
    """
    output = io.StringIO()
    
    for i in range(0, 256, 16):
        row_ints = sbox[i : i+16]
        row_str = " ".join([f"{val:02X}" for val in row_ints])
        output.write(row_str + "\n")
        
    output.seek(0)
    return output

def format_sbox_as_xlsx(sbox: List[int]) -> io.BytesIO:
    """
    Mengubah S-box menjadi format Excel (XLSX) Grid 16x16 dalam bentuk HEX (2 digit).
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "S-Box"

    for i in range(0, 256, 16):
        row_ints = sbox[i : i + 16]
        row_hex = [f"{val:02X}" for val in row_ints]
        ws.append(row_hex)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
